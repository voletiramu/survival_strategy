"""Create Excel trade report for 2026-03-11"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
green_font = Font(name='Arial', color='006100', size=10)
red_font = Font(name='Arial', color='9C0006', size=10)
data_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
title_font = Font(name='Arial', bold=True, size=14, color='1F4E79')
subtitle_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
warn_fill = PatternFill('solid', fgColor='FFF2CC')
warn_font = Font(name='Arial', color='9C6500', size=10, bold=True)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, pnl_col=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    if pnl_col:
        pnl_cell = ws.cell(row=row, column=pnl_col)
        val = pnl_cell.value
        if val is not None and isinstance(val, (int, float)):
            if val >= 0:
                pnl_cell.fill = green_fill
                pnl_cell.font = green_font
            else:
                pnl_cell.fill = red_fill
                pnl_cell.font = red_font

# ========== SUMMARY ==========
ws = wb.active
ws.title = 'Summary'
ws.sheet_properties.tabColor = '1F4E79'
ws['A1'] = 'ALGO TRADING - DAILY TRADE REPORT'
ws['A1'].font = title_font
ws.merge_cells('A1:H1')
ws['A2'] = 'Date: 11-March-2026 (Tuesday)'
ws['A2'].font = subtitle_font
ws.merge_cells('A2:H2')
ws['A3'] = 'Version: v10.2e | Branch: feature/angel-data-backtest'
ws['A3'].font = Font(name='Arial', italic=True, size=9, color='808080')

headers = ['Bot', 'Total Trades', 'Wins', 'Losses', 'Win Rate', 'Net PnL', 'Best Trade', 'Worst Trade']
for c, h in enumerate(headers, 1):
    ws.cell(row=5, column=c, value=h)
style_header(ws, 5, len(headers))

summary = [
    ['Equity', 15, 10, 5, 0.667, 114107.22, 23009.99, -570.01],
    ['Commodity', 7, 1, 6, 0.143, -11890.13, 155.49, -7141.98],
    ['Stock', 10, 3, 7, 0.300, 7413.68, 20300.50, -3728.42],
]
for r, row_data in enumerate(summary, 6):
    for c, val in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=val)
    style_data_row(ws, r, len(headers), pnl_col=6)
    ws.cell(row=r, column=5).number_format = '0.0%'
    for cc in [6, 7, 8]:
        ws.cell(row=r, column=cc).number_format = '#,##0.00'

ws.cell(row=9, column=1, value='TOTAL')
ws.cell(row=9, column=2, value='=SUM(B6:B8)')
ws.cell(row=9, column=3, value='=SUM(C6:C8)')
ws.cell(row=9, column=4, value='=SUM(D6:D8)')
ws.cell(row=9, column=5, value='=C9/B9')
ws.cell(row=9, column=6, value='=SUM(F6:F8)')
ws.cell(row=9, column=7, value='=MAX(G6:G8)')
ws.cell(row=9, column=8, value='=MIN(H6:H8)')
for c in range(1, 9):
    ws.cell(row=9, column=c).font = bold_font
    ws.cell(row=9, column=c).border = thin_border
    ws.cell(row=9, column=c).alignment = Alignment(horizontal='center')
ws.cell(row=9, column=5).number_format = '0.0%'
ws.cell(row=9, column=6).number_format = '#,##0.00'

ws['A11'] = 'KEY ISSUES IDENTIFIED'
ws['A11'].font = subtitle_font
ws.merge_cells('A11:H11')

issues = [
    ['CRITICAL', 'Stock bot uses Black-Scholes pricing (not real market data) - entry_oi=0, iv=25% hardcoded for ALL stock trades'],
    ['CRITICAL', 'SUNPHARMA 26-March expiry NOT found in instruments master -> fallback to BS model with phantom premiums'],
    ['HIGH', 'SUNPHARMA re-entered BUY_CE 7 times (same direction) - no direction-aware cooldown after repeated failures'],
    ['HIGH', 'GOLDM re-entered BUY_CE 5 times while gold was falling - CPR signal kept triggering bullish'],
    ['MEDIUM', 'Stock bot has NO signal logging file (unlike equity bot) - cannot audit signal history'],
    ['POSITIVE', 'Equity bot: high win rate on TRAILING_SL exits - TSL mechanism working well on bearish market'],
]
for r, (sev, desc) in enumerate(issues, 12):
    ws.cell(row=r, column=1, value=sev)
    ws.cell(row=r, column=2, value=desc)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    if sev == 'POSITIVE':
        ws.cell(row=r, column=1).font = Font(name='Arial', bold=True, color='006100', size=10)
    else:
        ws.cell(row=r, column=1).font = warn_font
    ws.cell(row=r, column=2).font = data_font
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

for c in range(1, 9):
    ws.column_dimensions[get_column_letter(c)].width = 18
ws.freeze_panes = 'A6'

# ========== EQUITY TRADES ==========
ws2 = wb.create_sheet('Equity Trades')
ws2.sheet_properties.tabColor = '2E75B6'
eq_headers = ['#', 'Symbol', 'Strategy', 'Signal', 'Strike', 'Expiry', 'Entry Time', 'Exit Time',
              'Hold (min)', 'Entry Prem', 'Exit Prem', 'Peak Prem', 'Num Lots', 'Lot Size',
              'Entry Spot', 'Entry IV', 'Entry OI', 'Exit Reason', 'Net PnL']
for c, h in enumerate(eq_headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(eq_headers))

eq_trades = [
    [1, 'NIFTY', 'CPR', 'BUY_PE_CPR', 24200, '17MAR2026', '09:37:10', '09:42:03', 5, 244.35, 248, 248, 1, 65, 24190.2, 34.1, 69313, 'BREAKOUT_FAIL', 116.78],
    [2, 'NIFTY', 'Gamma Blast', 'BUY_PE_GAMMA', 24200, '17MAR2026', '09:38:01', '09:42:56', 5, 244.35, 248, 248, 1, 65, 24182.85, 33.6, 69313, 'BREAKOUT_FAIL', 116.78],
    [3, 'BANKNIFTY', 'CPR', 'BUY_PE_CPR', 56400, '30MAR2026', '10:40:15', '10:45:11', 5, 996.75, 982.5, 996.75, 1, 30, 56384.6, 60.0, 1611, 'BREAKOUT_FAIL', -570.01],
    [4, 'SENSEX', 'CPR', 'BUY_PE_CPR', 78100, '12MAR2026', '09:30:01', '11:05:34', 96, 411.8, 701.6, 778.35, 4, 80, 78083.67, 18.1, 36059, 'TRAILING_SL_HIT', 23009.99],
    [5, 'SENSEX', 'PCR+VWAP', 'BUY_PE', 78100, '12MAR2026', '09:30:02', '11:05:35', 96, 411.8, 701.6, 778.35, 4, 80, 78083.67, 18.1, 36059, 'TRAILING_SL_HIT', 23009.99],
    [6, 'NIFTY', 'PCR+VWAP', 'BUY_PE', 24250, '17MAR2026', '09:32:27', '11:05:36', 93, 244.8, 331.3, 359.25, 1, 65, 24218.3, 32.6, 36612, 'TRAILING_SL_HIT', 5495.91],
    [7, 'SENSEX', 'Gamma Blast', 'BUY_PE_GAMMA', 78100, '12MAR2026', '09:33:03', '11:05:36', 93, 411.8, 701.6, 778.35, 3, 60, 77986.07, 15.9, 36059, 'TRAILING_SL_HIT', 17233.89],
    [8, 'NIFTY', 'CPR', 'BUY_PE_CPR', 24200, '17MAR2026', '09:53:03', '11:05:37', 73, 248.15, 306.75, 332.95, 1, 65, 24190.9, 34.7, 74097, 'TRAILING_SL_HIT', 3684.11],
    [9, 'NIFTY', 'Gamma Blast', 'BUY_PE_GAMMA', 24200, '17MAR2026', '09:53:22', '11:05:38', 72, 248.15, 306.75, 332.95, 1, 65, 24186.55, 34.4, 74097, 'TRAILING_SL_HIT', 3684.11],
    [10, 'SENSEX', 'CPR', 'BUY_PE_CPR', 77500, '12MAR2026', '11:15:46', '11:20:39', 5, 416.1, 416.1, 416.1, 4, 80, 77503.04, 18.8, 61403, 'BREAKOUT_FAIL', -148.49],
    [11, 'SENSEX', 'Gamma Blast', 'BUY_PE_GAMMA', 77500, '12MAR2026', '11:15:47', '11:20:39', 5, 416.1, 416.1, 416.1, 3, 60, 77503.04, 18.8, 61403, 'BREAKOUT_FAIL', -134.97],
    [12, 'BANKNIFTY', 'CPR', 'BUY_PE_CPR', 56400, '30MAR2026', '11:05:29', '12:18:21', 73, 955.15, 1151, 1239, 1, 30, 56445.2, 59.3, 2193, 'TRAILING_SL_HIT', 5727.92],
    [13, 'BANKNIFTY', 'Gamma Blast', 'BUY_PE_GAMMA', 56700, '30MAR2026', '09:30:00', '12:34:13', 184, 920.95, 1261.15, 1375, 2, 60, 56683.25, 55.1, 2651, 'TRAILING_SL_HIT', 20204.84],
    [14, 'BANKNIFTY', 'Gamma Blast', 'BUY_PE_GAMMA', 56100, '30MAR2026', '12:56:01', '13:00:48', 5, 1053.75, 1073.85, 1073.85, 2, 60, 56103.85, 64.3, 2316, 'BREAKOUT_FAIL', 1007.51],
    [15, 'BANKNIFTY', 'CPR', 'BUY_PE_CPR', 56700, '30MAR2026', '09:56:10', '13:27:40', 211, 983.1, 1377.25, 1480, 1, 30, 56593.15, 56.2, 3420, 'TRAILING_SL_HIT', 11668.87],
]
for r, row_data in enumerate(eq_trades, 2):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)
    style_data_row(ws2, r, len(eq_headers), pnl_col=19)
    ws2.cell(row=r, column=19).number_format = '#,##0.00'

r_total = len(eq_trades) + 2
ws2.cell(row=r_total, column=1, value='TOTAL')
ws2.cell(row=r_total, column=1).font = bold_font
ws2.cell(row=r_total, column=19, value='=SUM(S2:S16)')
ws2.cell(row=r_total, column=19).font = bold_font
ws2.cell(row=r_total, column=19).number_format = '#,##0.00'

col_widths_eq = [4, 12, 14, 16, 8, 12, 10, 10, 9, 11, 11, 11, 9, 9, 11, 9, 9, 22, 13]
for i, w in enumerate(col_widths_eq, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# ========== COMMODITY TRADES ==========
ws3 = wb.create_sheet('Commodity Trades')
ws3.sheet_properties.tabColor = 'ED7D31'
com_headers = ['#', 'Commodity', 'Strategy', 'Signal', 'Strike', 'Expiry', 'Entry Time', 'Exit Time',
               'Hold (min)', 'Entry Prem', 'Exit Prem', 'Peak Prem', 'Lots', 'Lot Size', 'Multiplier',
               'Entry Spot', 'Entry IV', 'Entry OI', 'Exit Reason', 'Net PnL']
for c, h in enumerate(com_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(com_headers))

com_trades = [
    [1, 'GOLDM', 'CPR', 'BUY_CE_CPR', 162900, '26MAR2026', '09:30:01', '09:34:56', 5, 4379, 4379, 4379, 2, 2, 10, 162751, 58.6, 800, 'BREAKOUT_FAIL', -148.70],
    [2, 'GOLDM', 'CPR', 'BUY_CE_CPR', 163000, '26MAR2026', '09:45:03', '09:49:57', 5, 4032, 4047, 4065.5, 2, 2, 10, 162880, 53.8, 111400, 'BREAKOUT_FAIL', 155.49],
    [3, 'GOLDM', 'CPR', 'BUY_CE_CPR', 162700, '26MAR2026', '10:59:28', '11:04:27', 5, 4284.5, 4284.5, 4284.5, 1, 1, 10, 162616, 57.0, 200, 'BREAKOUT_FAIL', -120.96],
    [4, 'GOLDM', 'CPR', 'BUY_CE_CPR', 162400, '26MAR2026', '11:00:57', '11:05:46', 5, 5368, 5368, 5368, 1, 1, 10, 162300, 71.6, 0, 'BREAKOUT_FAIL', -127.68],
    [5, 'GOLDM', 'CPR', 'BUY_CE_CPR', 163000, '26MAR2026', '10:00:04', '14:00:05', 240, 4040, 3690, 4074.5, 2, 2, 10, 162900, 53.8, 120500, 'TIME_EXIT', -7141.98],
    [6, 'CRUDEOILM', 'CPR', 'BUY_CE_CPR', 7650, '17MAR2026', '14:55:27', '15:34:03', 39, 947.55, 947.3, 1037.35, 5, 5, 10, 7593, 275.6, 6610, 'DIRECTION_FLIP', -136.27],
    [7, 'CRUDEOILM', 'Gamma Blast', 'BUY_PE_GAMMA', 7250, '17MAR2026', '19:31:04', '19:34:14', 3, 473.1, 387.85, 473.1, 5, 5, 10, 7201, 133.1, 9290, 'BREAKOUT_FAIL_REV', -4370.03],
]
for r, row_data in enumerate(com_trades, 2):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)
    style_data_row(ws3, r, len(com_headers), pnl_col=20)
    ws3.cell(row=r, column=20).number_format = '#,##0.00'

r_total = len(com_trades) + 2
ws3.cell(row=r_total, column=1, value='TOTAL')
ws3.cell(row=r_total, column=1).font = bold_font
ws3.cell(row=r_total, column=20, value='=SUM(T2:T8)')
ws3.cell(row=r_total, column=20).font = bold_font
ws3.cell(row=r_total, column=20).number_format = '#,##0.00'

col_widths_com = [4, 13, 14, 16, 8, 12, 10, 10, 9, 11, 11, 11, 6, 8, 10, 11, 9, 9, 20, 13]
for i, w in enumerate(col_widths_com, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

# ========== STOCK TRADES ==========
ws4 = wb.create_sheet('Stock Trades')
ws4.sheet_properties.tabColor = 'FF0000'
stk_headers = ['#', 'Symbol', 'Strategy', 'Signal', 'Strike', 'Expiry (Calc)', 'Entry Time', 'Exit Time',
               'Hold (min)', 'Entry Prem', 'Exit Prem', 'Peak Prem', 'Lot Size',
               'Entry Spot', 'Entry IV', 'Entry OI', 'Exit Reason', 'Raw PnL', 'Costs', 'Net PnL', 'Data Source']
for c, h in enumerate(stk_headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(stk_headers))

stk_trades = [
    [1, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1830, '26MAR (BS)', '09:30:20', '09:33:47', 3, 37.92, 28.75, 37.92, 350, 1827.1, 25.0, 0, 'BREAKOUT_FAIL_REV', -3209.5, 46.29, -3255.79, 'Black-Scholes'],
    [2, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1830, '26MAR (BS)', '09:34:20', '09:37:28', 3, 39.87, 29.35, 39.87, 350, 1830.8, 25.0, 0, 'BREAKOUT_FAIL_REV', -3682.0, 46.42, -3728.42, 'Black-Scholes'],
    [3, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1830, '26MAR (BS)', '09:47:53', '09:50:54', 3, 37.14, 27.0, 37.14, 350, 1825.6, 25.0, 0, 'BREAKOUT_FAIL_REV', -3549.0, 45.91, -3594.91, 'Black-Scholes'],
    [4, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1820, '26MAR (BS)', '10:01:16', '10:04:17', 3, 41.38, 31.3, 41.38, 350, 1824.0, 25.0, 0, 'BREAKOUT_FAIL_REV', -3528.0, 46.85, -3574.85, 'Black-Scholes'],
    [5, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1820, '26MAR (BS)', '10:14:42', '10:17:42', 3, 41.71, 32.95, 41.71, 350, 1824.6, 25.0, 0, 'BREAKOUT_FAIL_REV', -3066.0, 47.21, -3113.21, 'Black-Scholes'],
    [6, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1830, '26MAR (BS)', '10:27:04', '10:31:04', 4, 40.51, 32.3, 40.51, 350, 1832.0, 25.0, 0, 'BREAKOUT_FAIL_REV', -2873.5, 47.07, -2920.57, 'Black-Scholes'],
    [7, 'SBIN', 'CPR', 'BUY_PE_CPR', 1110, '26MAR (BS)', '09:30:18', '10:31:36', 61, 20.77, 25.9, 27.15, 750, 1110.4, 25.0, 0, 'CIRCUIT_BREAKER', 3847.5, 52.14, 3795.36, 'Black-Scholes'],
    [8, 'LAURUSLABS', 'CPR', 'BUY_CE_CPR', 1040, '26MAR (BS)', '09:30:21', '10:31:36', 61, 24.04, 48.0, 50.35, 850, 1043.0, 25.0, 0, 'CIRCUIT_BREAKER', 20366.0, 65.5, 20300.5, 'Black-Scholes'],
    [9, 'SUNPHARMA', 'CPR', 'BUY_CE_CPR', 1830, '26MAR (BS)', '10:41:47', '10:44:48', 3, 41.65, 31.95, 41.65, 350, 1834.1, 25.0, 0, 'BREAKOUT_FAIL_REV', -3395.0, 46.99, -3441.99, 'Black-Scholes'],
    [10, 'SBIN', 'CPR', 'BUY_PE_CPR', 1105, '26MAR (BS)', '10:44:51', '10:56:51', 12, 19.99, 24.35, 26.55, 750, 1106.9, 25.0, 0, 'TRAILING_SL_HIT', 3270.0, 51.41, 3218.59, 'Black-Scholes'],
]
for r, row_data in enumerate(stk_trades, 2):
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
    style_data_row(ws4, r, len(stk_headers), pnl_col=20)
    ws4.cell(row=r, column=20).number_format = '#,##0.00'
    ws4.cell(row=r, column=21).fill = warn_fill
    ws4.cell(row=r, column=21).font = warn_font
    oi_cell = ws4.cell(row=r, column=16)
    if oi_cell.value == 0:
        oi_cell.fill = red_fill
        oi_cell.font = red_font

r_total = len(stk_trades) + 2
ws4.cell(row=r_total, column=1, value='TOTAL')
ws4.cell(row=r_total, column=1).font = bold_font
ws4.cell(row=r_total, column=18, value='=SUM(R2:R11)')
ws4.cell(row=r_total, column=19, value='=SUM(S2:S11)')
ws4.cell(row=r_total, column=20, value='=SUM(T2:T11)')
for c in [18, 19, 20]:
    ws4.cell(row=r_total, column=c).font = bold_font
    ws4.cell(row=r_total, column=c).number_format = '#,##0.00'

ws4.cell(row=r_total+2, column=1, value='WARNING: ALL stock trades used Black-Scholes estimated premiums (NOT real market data)')
ws4.cell(row=r_total+2, column=1).font = Font(name='Arial', bold=True, color='FF0000', size=11)
ws4.merge_cells(start_row=r_total+2, start_column=1, end_row=r_total+2, end_column=21)
ws4.cell(row=r_total+3, column=1, value='Root cause: SUNPHARMA 26-March expiry not found in Angel instruments master -> fallback BS pricing with hardcoded IV=25%')
ws4.cell(row=r_total+3, column=1).font = Font(name='Arial', italic=True, color='808080', size=9)
ws4.merge_cells(start_row=r_total+3, start_column=1, end_row=r_total+3, end_column=21)

col_widths_stk = [4, 13, 8, 16, 7, 13, 10, 10, 9, 11, 10, 10, 8, 11, 9, 9, 20, 10, 8, 13, 14]
for i, w in enumerate(col_widths_stk, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A2'

output_path = r'C:\Users\Ram\Data\algo_trading\trades_20260311.xlsx'
wb.save(output_path)
print('Excel saved to', output_path)
