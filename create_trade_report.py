from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
header_fill = PatternFill('solid', fgColor='4472C4')
header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
data_font = Font(name='Arial', size=10)
bold_font = Font(bold=True, name='Arial', size=10)
money_fmt = '#,##0.00'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_data_row(ws, row, cols, pnl_col):
    pnl = ws.cell(row=row, column=pnl_col).value
    fill = green_fill if pnl and pnl > 0 else red_fill
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = fill

# ============ SHEET 1: EQUITY TRADES ============
ws1 = wb.active
ws1.title = "Equity Trades"

eq_headers = ['#', 'Symbol', 'Signal Type', 'Strike', 'Entry Time', 'Exit Time',
              'Hold (min)', 'Entry Premium', 'Exit Premium', 'Peak Premium',
              'Lots', 'Lot Size', 'PnL (Rs)', 'Exit Reason', 'Quality Score',
              'Delta', 'IV%', 'Entry Spot', 'DTE', 'Post-Mortem Analysis']

ws1.append(eq_headers)
style_header(ws1, 1, len(eq_headers))

equity_trades = [
    [1, 'NIFTY', 'BUY_PE_CPR', 23800, '12:04:24', '12:09:21', 5, 274.90, 276.05, 276.05, 3, 195, 42.49,
     'BREAKOUT_FAIL', 85, -0.526, 50.2, 23746, 1,
     'SMALL WIN but exited too early. Spot moved -25pts in our favor but premium barely moved (+0.4%). '
     'BREAKOUT_FAIL timer (5min/2% threshold) killed a valid trade. NIFTY later dropped to 23650 — '
     'this PE would have hit target. ISSUE: 5min breakout timer too aggressive for options that need time to react.'],

    [2, 'BANKNIFTY', 'BUY_PE_CPR', 55500, '12:04:25', '12:09:22', 5, 1190, 1200, 1200, 2, 60, 388.90,
     'BREAKOUT_FAIL', 83, -0.533, 90.4, 55210, 1,
     'SMALL WIN. Spot dropped -66pts, premium gained only Rs 10 (+0.8%). Exited by BREAKOUT_FAIL timer. '
     'High IV=90.4% means premium was already inflated and less responsive to spot moves. '
     'ISSUE: Same 5min timer. Also, very high IV entry — premium paid was expensive.'],

    [3, 'BANKNIFTY', 'SELL_PE_CPR', 55000, '12:04:26', '12:09:22', 5, 994.95, 1000.90, 994.95, 1, 30, -321.49,
     'BREAKOUT_FAIL', 55, -0.458, 95.7, 55210, 1,
     'LOSS on SELL trade. Sold PE hoping for mean reversion at S3, but spot kept falling (-66pts). '
     'Premium rose against us. Quality score was LOWEST at 55 — should have been filtered. '
     'ISSUE: Selling options in HIGH VIX (21.5) + SIDEWAYS regime is dangerous. '
     'SELL signals need higher quality threshold (70+). Also contradicted the BUY_PE signal on same symbol.'],

    [4, 'SENSEX', 'BUY_PE_CPR', 76500, '12:04:27', '12:22:30', 18, 230.20, 193.40, 254.30, 4, 80, -3065.02,
     'TRAILING_SL_HIT', 100, -0.528, 13.2, 76447, 1,
     'BIG LOSS. Premium peaked at 254.30 (+10.5%) then collapsed to 193.40. TSL set at 196.79 was hit. '
     'Spot went from 76447 down to ~76380 (our favor) but then reversed sharply upward. '
     'ISSUE: 0-DTE SENSEX expiry — extreme theta decay. IV=13.2% (very low) means tiny moves kill premium. '
     'The TSL locked in a loss because premium decayed even as spot moved our way initially. '
     'CRITICAL: Trading 0-DTE with low IV is a theta trap. Bot should avoid 0-DTE entries.'],

    [5, 'BANKNIFTY', 'BUY_PE_CPR', 55500, '12:19:41', '12:24:26', 5, 1176.95, 1147.90, 1176.95, 1, 30, -1022.29,
     'BREAKOUT_FAIL', 83, -0.537, 88.4, 55193, 1,
     'LOSS on re-entry. This was REENTRY #2 after the first BANKNIFTY PE closed. Entered at 1176.95 but '
     'premium immediately dropped to 1147.90 (-2.5%). Spot didn\'t move enough to offset theta. '
     'ISSUE: Re-entering the same trade within 15min without new confirmation. Cooldown was bypassed. '
     'The breakout momentum was already exhausted from the first entry.'],

    [6, 'SENSEX', 'BUY_PE_CPR', 76400, '12:42:35', '12:45:39', 3, 220.05, 186.95, 220.05, 4, 80, -2768.03,
     'BREAKOUT_FAIL_REVERSE', 100, -0.523, 12.8, 76355, 1,
     'BIG LOSS. SENSEX 2nd attempt. Entry at 220.05, immediately reversed to 186.95 (-15%). '
     'BREAKOUT_FAIL_REVERSE triggered in just 3 minutes. Spot reversed against the bearish thesis. '
     'ISSUE: Same 0-DTE problem. IV=12.8% — premium melts instantly on any reversal. '
     'Quality=100 was misleading because it didn\'t penalize 0-DTE expiry or low IV. '
     'Bot re-entered SENSEX PE despite previous -3065 loss on same thesis.'],

    [7, 'SENSEX', 'BUY_PE_CPR', 76400, '13:25:57', '13:29:03', 3, 167.65, 134.05, 167.65, 4, 80, -2801.17,
     'BREAKOUT_FAIL_REVERSE', 100, -0.460, 11.9, 76432, 1,
     'BIG LOSS. SENSEX 3rd attempt! Entry at 167.65, crashed to 134.05 (-20%) in 3 minutes. '
     'By now SENSEX had lost -5834 on 3 failed PE trades on the SAME thesis. '
     'ISSUE: NO per-symbol daily loss limit. Bot kept re-entering SENSEX PE despite repeated failures. '
     'IV dropped to 11.9% — each re-entry had worse theta decay. '
     'CRITICAL: Need per-symbol max loss cap (e.g., -3000/day) and max re-entries per symbol (2).'],

    [8, 'NIFTY', 'BUY_PE_CPR', 23750, '12:19:40', '13:29:28', 70, 247.35, 225.75, 273.95, 3, 195, -4380.04,
     'TRAILING_SL_HIT', 90, -0.510, 47.6, 23723, 1,
     'BIGGEST LOSS. Held for 70 minutes. Premium peaked at 273.95 (+10.8%) but TSL at 225.86 was hit. '
     'Spot was choppy — dropped to favor us, then reversed. The TSL was too tight (trail=30% in SIDEWAYS). '
     'Peak unrealized was +Rs 15,561 but ended at -Rs 4,380. '
     'ISSUE: TSL trailed 30% of peak which means even a small pullback after gains triggers exit. '
     'For a 70min hold, this is far too aggressive. The trade WAS right directionally but got shaken out.'],

    [9, 'NIFTY', 'BUY_PE_CPR', 23850, '13:49:30', '15:20:07', 91, 257.45, 341.05, 355.90, 3, 195, 16107.68,
     'EOD_FORCE_CLOSE', 90, -0.538, 45.4, 23784, 1,
     'BIG WINNER! Entry at 257.45, hit target extension, peak at 355.90 (+38.2%). Breakeven locked. '
     'TSL at 326.37 held through pullbacks. Exited at EOD at 341.05. This is how the strategy SHOULD work. '
     'NIFTY dropped sharply in the afternoon, PE premium exploded. '
     'KEY: This was the 3rd NIFTY PE attempt — persistence paid off but only because market finally moved big.'],
]

for row_data in equity_trades:
    ws1.append(row_data)
    style_data_row(ws1, ws1.max_row, len(eq_headers), 13)

# Summary row
ws1.append([])
ws1.append(['', '', '', '', '', '', '', '', '', '', '', '', '=SUM(M2:M10)', 'NET EQUITY PnL'])
ws1.cell(row=ws1.max_row, column=13).number_format = money_fmt
ws1.cell(row=ws1.max_row, column=13).font = Font(bold=True, name='Arial', size=11)
ws1.cell(row=ws1.max_row, column=14).font = bold_font

# Column widths
eq_widths = [4, 12, 14, 10, 10, 10, 8, 12, 12, 12, 6, 8, 12, 20, 8, 8, 6, 10, 5, 80]
for i, w in enumerate(eq_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Format number columns
for row in range(2, 11):
    for col in [8, 9, 10, 13]:
        ws1.cell(row=row, column=col).number_format = money_fmt

# ============ SHEET 2: COMMODITY TRADES ============
ws2 = wb.create_sheet("Commodity Trades")

com_headers = ['#', 'Commodity', 'Signal Type', 'Strike', 'Entry Time', 'Exit Time',
               'Hold (min)', 'Entry Premium', 'Exit Premium', 'Peak Premium',
               'Lots', 'Multiplier', 'PnL (Rs)', 'Exit Reason', 'Quality Score',
               'IV%', 'Entry Spot', 'DTE', 'Post-Mortem Analysis']

ws2.append(com_headers)
style_header(ws2, 1, len(com_headers))

commodity_trades = [
    [1, 'GOLDM', 'BUY_CE_CPR', 162000, '13:01:39', '13:06:28', 5, 3573.50, 3510.50, 3573.50, 2, 10, -1398.26,
     'BREAKOUT_FAIL', 85, 47.9, 161915, 5,
     'LOSS. First commodity entry. Premium dropped immediately. Breakout didn\'t sustain — '
     'GOLDM was in SIDEWAYS regime. 5min BREAKOUT_FAIL timer killed it. Gold was choppy around 161900.'],

    [2, 'SILVERM', 'BUY_CE_GAMMA', 280000, '13:03:22', '13:08:15', 5, 13530, 13650.50, 13700, 1, 5, 465.94,
     'BREAKOUT_FAIL', 85, 108.4, 278799, 5,
     'SMALL WIN. Gamma Blast strategy. Premium peaked at 13700 (+1.3%) but BREAKOUT_FAIL exited early. '
     'Silver was surging (+6658 body) but the 5min timer exited before the move could develop.'],

    [3, 'SILVERM', 'BUY_CE_CPR', 280000, '13:05:13', '13:10:08', 5, 13621, 13564.50, 13650.50, 1, 5, -419.03,
     'BREAKOUT_FAIL', 85, 108.9, 278855, 5,
     'LOSS. Duplicate SILVERM CE entry alongside the Gamma Blast trade. Same strike, same thesis. '
     'ISSUE: CPR and Gamma Blast taking same trade = double exposure.'],

    [4, 'SILVERM', 'BUY_CE_CPR', 280000, '13:30:12', '13:35:12', 5, 13900, 13910, 13988, 1, 5, -87.51,
     'BREAKOUT_FAIL', 85, 110.1, 279098, 5,
     'TINY LOSS. Re-entry on SILVERM CE. Premium barely moved. Brokerage costs ate the small gain.'],

    [5, 'SILVERM', 'BUY_CE_GAMMA', 280000, '13:30:13', '13:35:12', 5, 13900, 13910, 13988, 1, 5, -87.51,
     'BREAKOUT_FAIL', 85, 110.1, 279098, 5,
     'DUPLICATE of Trade #4. Same strike, same time, same exit. CPR+Gamma Blast doubling up again.'],

    [6, 'CRUDEOILM', 'SELL_CE_CPR', 7900, '13:36:09', '13:41:09', 5, 896.15, 900, 896.15, 5, 10, -314.73,
     'BREAKOUT_FAIL', 70, 272, 7714, 5,
     'LOSS. Sold CE for mean reversion at R3. But crude was in a massive uptrend (+354pts from open). '
     'IV=272% — extremely volatile. Selling into a strong trend with sky-high IV is very risky. '
     'Quality=70 was borderline. ISSUE: Sell signals in strong trends should be blocked.'],

    [7, 'GOLDM', 'BUY_CE_CPR', 162000, '13:50:49', '13:55:37', 5, 3610, 3516, 3610, 2, 10, -2018.49,
     'BREAKOUT_FAIL', 85, 48.1, 161950, 5,
     'BIGGEST COMMODITY LOSS. Premium dropped -94pts immediately. Gold was not breaking out — '
     'it was rangebound around 161900-162000. Narrow CPR (0.158%) gave false breakout signal.'],

    [8, 'GOLDM', 'BUY_CE_CPR', 162000, '13:55:54', '14:00:44', 5, 3516, 3529, 3549, 2, 10, 121.91,
     'BREAKOUT_FAIL', 85, 47.4, 161866, 5,
     'TINY WIN. 3rd GOLDM CE attempt. Barely profitable after costs. Gold still ranging.'],

    [9, 'SILVERM', 'BUY_CE_GAMMA', 280000, '13:55:56', '14:00:45', 5, 13694, 13650, 13709.50, 1, 5, -356.77,
     'BREAKOUT_FAIL', 85, 109.8, 278768, 5,
     'LOSS. Another SILVERM Gamma Blast. Same 280000 strike for the 4th time. Silver was cooling off '
     'from its initial surge. Re-entering exhausted breakouts.'],

    [10, 'GOLDM', 'BUY_CE_CPR', 162000, '14:23:54', '14:28:44', 5, 3492, 3534.50, 3526.50, 2, 10, 711.99,
     'BREAKOUT_FAIL', 85, 47.3, 161844, 5,
     'WIN. 4th GOLDM CE. Finally caught a small move. But still exited by 5min timer.'],

    [11, 'GOLDM', 'BUY_CE_CPR', 162000, '14:30:46', '14:35:40', 5, 3540, 3560.50, 3586.50, 2, 10, 271.55,
     'BREAKOUT_FAIL', 85, 47.8, 161862, 5,
     'SMALL WIN. 5th GOLDM CE. Pattern: keep re-entering same trade, small wins/losses each time. '
     'Net effect is churning with brokerage eating profits.'],

    [12, 'GOLDM', 'BUY_CE_CPR', 162000, '14:35:57', '16:24:11', 109, 3560.50, 3562.50, 3735, 2, 10, -98.56,
     'SIGNAL_WEAK_EXIT', 85, 47.9, 161889, 5,
     'LOSS despite peaking at 3735 (+4.9%). Held 109min but exited at 3562.50 — gave back all gains. '
     'SIGNAL_WEAK_EXIT triggered. Gold peaked at ~162200 then fell back. '
     'TSL should have locked breakeven at 3735 peak.'],
]

for row_data in commodity_trades:
    ws2.append(row_data)
    style_data_row(ws2, ws2.max_row, len(com_headers), 13)

ws2.append([])
ws2.append(['', '', '', '', '', '', '', '', '', '', '', '', '=SUM(M2:M13)', 'NET COMMODITY PnL'])
ws2.cell(row=ws2.max_row, column=13).number_format = money_fmt
ws2.cell(row=ws2.max_row, column=13).font = Font(bold=True, name='Arial', size=11)
ws2.cell(row=ws2.max_row, column=14).font = bold_font

com_widths = [4, 12, 14, 10, 10, 10, 8, 12, 12, 12, 6, 8, 12, 20, 8, 6, 10, 5, 80]
for i, w in enumerate(com_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for row in range(2, 14):
    for col in [8, 9, 10, 13]:
        ws2.cell(row=row, column=col).number_format = money_fmt

# ============ SHEET 3: SUMMARY & KEY FINDINGS ============
ws3 = wb.create_sheet("Summary")

title_font = Font(bold=True, name='Arial', size=14, color='1F4E79')
section_font = Font(bold=True, name='Arial', size=12, color='2E75B6')
ws3.column_dimensions['A'].width = 35
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 60

ws3['A1'] = 'TRADE ANALYSIS — 12 March 2026'
ws3['A1'].font = title_font

ws3['A3'] = 'OVERALL PERFORMANCE'
ws3['A3'].font = section_font

labels = [
    ('Total Equity PnL', "='Equity Trades'!M11", money_fmt),
    ('Total Commodity PnL', "='Commodity Trades'!M14", money_fmt),
    ('Combined PnL', "='Equity Trades'!M11+'Commodity Trades'!M14", money_fmt),
    ('Equity Trades', '9 (3W / 6L)', None),
    ('Commodity Trades', '12 (4W / 8L)', None),
    ('Equity Win Rate', '33.3%', None),
    ('Commodity Win Rate', '33.3%', None),
    ('Biggest Winner', 'NIFTY 23850PE +Rs 16,108', None),
    ('Biggest Loser', 'NIFTY 23750PE -Rs 4,380', None),
    ('Most Traded (Equity)', 'SENSEX (3 trades, -Rs 8,634)', None),
    ('Most Traded (Commodity)', 'GOLDM (6 trades, -Rs 2,409)', None),
]

for i, (label, val, fmt) in enumerate(labels, 4):
    ws3.cell(row=i, column=1, value=label).font = bold_font
    ws3.cell(row=i, column=2, value=val).font = data_font
    if fmt:
        ws3.cell(row=i, column=2).number_format = fmt

ws3['A17'] = 'KEY ISSUES IDENTIFIED'
ws3['A17'].font = section_font

issues = [
    ('1. BREAKOUT_FAIL Timer Too Aggressive',
     '14 of 21 trades exited by BREAKOUT_FAIL (5min/2% threshold). Most trades never got a chance to develop. '
     'Options need more time than spot to react — a 5min window with 2% threshold is killing valid trades.'),
    ('2. No Per-Symbol Daily Loss Limit',
     'SENSEX lost Rs 8,634 across 3 trades on the SAME bearish thesis. Bot kept re-entering despite repeated failures. '
     'Need a per-symbol daily loss cap (e.g., Rs 3,000) and max 2 re-entries per symbol per day.'),
    ('3. 0-DTE Theta Trap (SENSEX)',
     'All 3 SENSEX trades were 0-DTE with IV=11-13%. Theta decay was so extreme that even favorable spot moves '
     'couldn\'t save the premium. Bot must avoid 0-DTE entries or significantly reduce position size.'),
    ('4. Duplicate Strategy Exposure (Commodities)',
     'CPR and Gamma Blast took identical trades (same strike, same time) on SILVERM — doubling risk. '
     'Need deduplication: if CPR and Gamma Blast signal the same strike, take only ONE position.'),
    ('5. SELL Signals in Trending Markets',
     'BANKNIFTY SELL_PE (quality=55) and CRUDEOILM SELL_CE (IV=272%) were mean-reversion bets against strong trends. '
     'Sell signals need: higher quality threshold (70+), trend filter, and should be blocked in HIGH VIX regime.'),
    ('6. Trailing SL Too Tight in SIDEWAYS Regime',
     'NIFTY 23750PE peaked at +Rs 15,561 unrealized but TSL (30% trail) took it out at -Rs 4,380. '
     'For SIDEWAYS regime, 30% trail is too tight — consider 40-50% or using time-based trailing.'),
    ('7. Excessive Churning on Same Symbol',
     'GOLDM had 6 trades on 162000CE, SILVERM had 5 trades on 280000CE. Most were tiny gains/losses eaten by brokerage. '
     'Need a cooldown period between same-symbol re-entries (at least 30min for commodities).'),
]

row = 18
for title, desc in issues:
    ws3.cell(row=row, column=1, value=title).font = bold_font
    ws3.cell(row=row, column=2, value=desc).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 2

ws3[f'A{row+1}'] = 'RECOMMENDED FIXES'
ws3[f'A{row+1}'].font = section_font
row += 2

fixes = [
    'Increase BREAKOUT_FAIL timer from 5min to 10-15min, or raise threshold from 2% to 5%',
    'Add per-symbol daily loss limit: Rs 3,000 for equity, Rs 2,000 for commodity',
    'Block 0-DTE entries entirely OR require DTE >= 1 for all entries',
    'Deduplicate CPR + Gamma Blast signals — if same strike, take only one position',
    'Raise SELL signal quality threshold from 50 to 70, block sells in HIGH VIX regime',
    'Widen TSL trail to 40-50% in SIDEWAYS regime, or use time-based trailing after 30min',
    'Add 30min cooldown between same-symbol re-entries for commodities',
    'Cap max entries per symbol per day: 2 for equity, 3 for commodity',
]

for i, fix in enumerate(fixes, 1):
    ws3.cell(row=row, column=1, value=f'Fix {i}').font = bold_font
    ws3.cell(row=row, column=2, value=fix).font = data_font
    ws3.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
    ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    row += 1

wb.save('C:/Users/Ram/Data/algo_trading/trade_analysis_20260312.xlsx')
print("Saved trade_analysis_20260312.xlsx")
